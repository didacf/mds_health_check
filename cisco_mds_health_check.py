#!/usr/bin/env python3
"""
Cisco MDS 9000 Series Health Check Script
==========================================
Connects to one or more MDS switches via SSH, runs diagnostic commands,
parses the output, and generates a color-coded health report.

Requirements:
    pip install netmiko

Usage:
    python cisco_mds_health_check.py                        # interactive prompt
    python cisco_mds_health_check.py -t 10.1.1.1            # single switch
    python cisco_mds_health_check.py -f switches.txt        # file with one IP per line
    python cisco_mds_health_check.py -t 10.1.1.1 -u admin   # specify username
    python cisco_mds_health_check.py -t 10.1.1.1 --json     # JSON output

    # Compare zoning between two fabrics:
    python cisco_mds_health_check.py --fabric-a 10.1.1.1 --fabric-b 10.2.2.1
"""

import argparse
import getpass
import json
import re
import sys
import os
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional

try:
    from netmiko import ConnectHandler, NetmikoTimeoutException, NetmikoAuthenticationException
except ImportError:
    print("ERROR: netmiko is required. Install with: pip install netmiko")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────
# Data models
# ─────────────────────────────────────────────

@dataclass
class Finding:
    """A single health-check finding."""
    module: str
    severity: str          # CRITICAL, WARNING, INFO, OK
    message: str
    detail: str = ""

@dataclass
class SwitchReport:
    """Health report for one switch."""
    hostname: str = ""
    ip: str = ""
    model: str = ""
    serial: str = ""
    nxos_version: str = ""
    uptime: str = ""
    timestamp: str = ""
    findings: list = field(default_factory=list)

    @property
    def critical_count(self):
        return sum(1 for f in self.findings if f.severity == "CRITICAL")

    @property
    def warning_count(self):
        return sum(1 for f in self.findings if f.severity == "WARNING")

    @property
    def overall_status(self):
        if self.critical_count > 0:
            return "CRITICAL"
        if self.warning_count > 0:
            return "WARNING"
        return "HEALTHY"


# ─────────────────────────────────────────────
# Terminal colors
# ─────────────────────────────────────────────

class Colors:
    RED = "\033[91m"
    YELLOW = "\033[93m"
    GREEN = "\033[92m"
    CYAN = "\033[96m"
    BOLD = "\033[1m"
    DIM = "\033[2m"
    RESET = "\033[0m"

    @staticmethod
    def severity(sev: str) -> str:
        return {
            "CRITICAL": f"{Colors.RED}{Colors.BOLD}CRITICAL{Colors.RESET}",
            "WARNING":  f"{Colors.YELLOW}WARNING{Colors.RESET}",
            "INFO":     f"{Colors.CYAN}INFO{Colors.RESET}",
            "OK":       f"{Colors.GREEN}OK{Colors.RESET}",
        }.get(sev, sev)


# ─────────────────────────────────────────────
# Command runner
# ─────────────────────────────────────────────

HEALTH_COMMANDS = [
    "show version",
    "show module",
    "show hardware",
    "show inventory",
    "show interface brief",
    "show interface",
    "show interface transceiver details",
    "show environment",
    "show logging last 50",
    "show system uptime",
    "show running-config",
    "show clock",
    "show flogi database",
    "show fcns database local vsan 1-4093",
    "show port-channel summary",
    "show port-channel database",
    "show vsan",
    "show vsan membership",
    "show fcdomain vsan 1-4093",
    "show zone status",
    "show zoneset active",
    "show process cpu sort",
    "show system resources",
    "show accounting log last 25",
    "show device-alias database",
    "show device-alias status",
    "show zone status vsan 1-4093",
    "show zoneset active vsan 1-4093",
    "show zone name",
    "show hardware internal errors",
    "show hardware internal fcmac acl-sobp resource-usage",
    "show interface counters detailed",
    "show port-channel usage",
    "show port-channel consistency",
]


def connect_switch(ip: str, username: str, password: str, timeout: int = 30) -> Optional[object]:
    """Establish SSH connection to an MDS switch."""
    device = {
        "device_type": "cisco_nxos",
        "host": ip,
        "username": username,
        "password": password,
        "timeout": timeout,
        "conn_timeout": timeout,
    }
    try:
        print(f"  Connecting to {ip} ...")
        conn = ConnectHandler(**device)
        print(f"  {Colors.GREEN}Connected.{Colors.RESET}")
        return conn
    except NetmikoAuthenticationException:
        print(f"  {Colors.RED}Authentication failed for {ip}.{Colors.RESET}")
    except NetmikoTimeoutException:
        print(f"  {Colors.RED}Connection timed out for {ip}.{Colors.RESET}")
    except Exception as e:
        print(f"  {Colors.RED}Connection error for {ip}: {e}{Colors.RESET}")
    return None


def collect_outputs(conn, commands: list) -> dict:
    """Run a list of CLI commands and return {command: output}."""
    results = {}
    for cmd in commands:
        try:
            output = conn.send_command(cmd, read_timeout=60)
            results[cmd] = output
        except Exception:
            results[cmd] = ""
    return results


# ─────────────────────────────────────────────
# Health check modules
# ─────────────────────────────────────────────

def check_version(outputs: dict, report: SwitchReport):
    """Parse show version for basic info and NX-OS release."""
    raw = outputs.get("show version", "")
    if not raw:
        report.findings.append(Finding("NX-OS Version", "WARNING", "Could not retrieve show version output."))
        return

    # Extract hostname
    m = re.search(r"Device name:\s*(\S+)", raw)
    if m:
        report.hostname = m.group(1)

    # Extract NX-OS version
    m = re.search(r"(?:system|NXOS):\s+version\s+(\S+)", raw, re.I)
    if not m:
        m = re.search(r"(?:kickstart|system) image file is:\s+\S+\.(\S+)", raw, re.I)
    if m:
        report.nxos_version = m.group(1)

    # Extract chassis / model
    m = re.search(r"cisco\s+(MDS\s+\S+)", raw, re.I)
    if m:
        report.model = m.group(1)

    # Check for known old/critical releases (example baseline)
    ver = report.nxos_version
    if ver:
        report.findings.append(Finding("NX-OS Version", "INFO",
            f"Running NX-OS version {ver}.",
            "Verify this is a Cisco-recommended release at cisco.com/go/mds."))
    else:
        report.findings.append(Finding("NX-OS Version", "WARNING", "Could not determine NX-OS version."))


def check_uptime(outputs: dict, report: SwitchReport):
    """Check system uptime."""
    raw = outputs.get("show system uptime", "")
    m = re.search(r"System uptime:\s+(.+)", raw)
    if m:
        report.uptime = m.group(1).strip()
        # Warn if uptime < 1 day (recent reboot)
        if re.search(r"0\s+day", report.uptime):
            report.findings.append(Finding("Uptime", "WARNING",
                f"Switch rebooted recently — uptime: {report.uptime}"))
        else:
            report.findings.append(Finding("Uptime", "OK",
                f"System uptime: {report.uptime}"))
    else:
        report.findings.append(Finding("Uptime", "INFO", "Could not parse system uptime."))


def check_modules(outputs: dict, report: SwitchReport):
    """Check module status from show module."""
    raw = outputs.get("show module", "")
    if not raw:
        return

    bad_modules = []
    for line in raw.splitlines():
        # Module lines typically start with a number
        m = re.match(r"\s*(\d+)\s+\d+\s+(.+?)\s+(ok|active|ha-standby|standby|powered-down|powered-up|failure|err)\s*", line, re.I)
        if m:
            mod_num, mod_desc, status = m.group(1), m.group(2).strip(), m.group(3).lower()
            if status in ("ok", "active", "ha-standby", "standby", "powered-up"):
                continue
            bad_modules.append(f"Module {mod_num} ({mod_desc}): status={status}")

    if bad_modules:
        report.findings.append(Finding("Module Status", "CRITICAL",
            f"{len(bad_modules)} module(s) in abnormal state.",
            "\n".join(bad_modules)))
    else:
        report.findings.append(Finding("Module Status", "OK", "All modules are in normal state."))


def check_environment(outputs: dict, report: SwitchReport):
    """Check power supply and fan status from show environment."""
    raw = outputs.get("show environment", "")
    if not raw:
        report.findings.append(Finding("Environment", "INFO", "No environment data available."))
        return

    issues = []

    # Check for fan failures
    for line in raw.splitlines():
        line_lower = line.lower()
        if any(kw in line_lower for kw in ("fail", "absent", "shutdown", "fault")):
            if any(kw in line_lower for kw in ("fan", "power", "ps", "temperature")):
                issues.append(line.strip())

    # Check temperature warnings
    temp_matches = re.findall(r"(\d+)\s*C\b", raw)
    for temp_str in temp_matches:
        temp = int(temp_str)
        if temp > 65:
            issues.append(f"High temperature detected: {temp}°C")

    if issues:
        report.findings.append(Finding("Environment", "CRITICAL" if len(issues) > 2 else "WARNING",
            f"{len(issues)} environmental issue(s) detected.",
            "\n".join(issues[:10])))
    else:
        report.findings.append(Finding("Environment", "OK", "Power supplies, fans, and temperature normal."))


def check_interfaces(outputs: dict, report: SwitchReport):
    """Check interface state summary."""
    raw_brief = outputs.get("show interface brief", "")

    up_count = 0
    down_count = 0

    for line in raw_brief.splitlines():
        m = re.match(r"\s*(fc\d+/\d+)\s+", line, re.I)
        if m:
            if "up" in line.lower().split()[1:3]:
                up_count += 1
            elif "down" in line.lower():
                down_count += 1

    report.findings.append(Finding("Interfaces", "INFO",
        f"FC interfaces — Up: {up_count}, Down/notconnect: {down_count}"))


# Error counter definitions: (regex_pattern, display_name, warn_threshold, crit_threshold)
_ERROR_COUNTERS = [
    (r"(\d+)\s+input error",                "input errors",             1,   100),
    (r"(\d+)\s+crc",                         "CRC errors",              1,   50),
    (r"(\d+)\s+invalid crc",                 "invalid CRC",             1,   50),
    (r"(\d+)\s+invalid transmission word",   "invalid tx words",        1,   100),
    (r"(\d+)\s+invalid ordered set",         "invalid ordered sets",    1,   100),
    (r"(\d+)\s+link failure",                "link failures",           1,   10),
    (r"(\d+)\s+signal loss",                 "signal losses",           1,   10),
    (r"(\d+)\s+sync loss",                   "sync losses",             1,   10),
    (r"(\d+)\s+link reset",                  "link resets",             5,   50),
    (r"(\d+)\s+link reset received",         "link resets rx",          5,   50),
    (r"(\d+)\s+timeout discard",             "timeout discards",        1,   100),
    (r"(\d+)\s+credit loss",                 "credit losses",           1,   50),
    (r"(\d+)\s+fec corrected",               "FEC corrected blocks",    1000, 50000),
    (r"(\d+)\s+fec uncorrected",             "FEC uncorrected blocks",  1,   10),
    (r"(\d+)\s+non.?corrected",              "non-corrected errors",    1,   10),
    (r"(\d+)\s+delimiter err",               "delimiter errors",        1,   50),
    (r"(\d+)\s+address id err",              "address ID errors",       1,   50),
]


def check_interface_errors(outputs: dict, report: SwitchReport):
    """Detailed per-interface error counter analysis: CRC, invalid words,
    FEC corrected/uncorrected, link failures, signal/sync loss, credit loss."""
    raw_full = outputs.get("show interface", "")
    raw_detailed = outputs.get("show interface counters detailed", "")
    combined = raw_full + "\n" + raw_detailed

    if not combined.strip():
        report.findings.append(Finding("Interface Errors", "INFO",
            "No interface counter data available."))
        return

    # Parse per-interface error counters
    current_intf = ""
    # {intf: [(counter_name, count, severity)]}
    intf_errors: dict = {}

    for line in combined.splitlines():
        im = re.match(r"^(fc\d+/\d+|port-channel\d+)\s+is\s+", line, re.I)
        if im:
            current_intf = im.group(1)
            continue
        # Also catch "fcX/Y" at start of counters-detailed blocks
        im2 = re.match(r"^(fc\d+/\d+|port-channel\d+)\s*$", line.strip(), re.I)
        if im2:
            current_intf = im2.group(1)
            continue
        if not current_intf:
            continue

        for pattern, name, warn_thresh, crit_thresh in _ERROR_COUNTERS:
            cm = re.search(pattern, line, re.I)
            if cm:
                count = int(cm.group(1))
                if count > 0:
                    if count >= crit_thresh:
                        sev = "CRITICAL"
                    elif count >= warn_thresh:
                        sev = "WARNING"
                    else:
                        sev = "INFO"
                    intf_errors.setdefault(current_intf, []).append((name, count, sev))

    # Separate findings by severity
    critical_items = []
    warning_items = []
    info_items = []

    for intf, errors in sorted(intf_errors.items()):
        for name, count, sev in errors:
            entry = f"{intf}: {count:,} {name}"
            if sev == "CRITICAL":
                critical_items.append(entry)
            elif sev == "WARNING":
                warning_items.append(entry)
            else:
                info_items.append(entry)

    if critical_items:
        report.findings.append(Finding("Interface Errors", "CRITICAL",
            f"{len(critical_items)} critical error counter(s) detected.",
            "\n".join(critical_items[:25])))
    if warning_items:
        report.findings.append(Finding("Interface Errors", "WARNING",
            f"{len(warning_items)} elevated error counter(s) detected.",
            "\n".join(warning_items[:25])))
    if info_items:
        report.findings.append(Finding("Interface Errors", "INFO",
            f"{len(info_items)} low-count error counter(s) (monitor for growth).",
            "\n".join(info_items[:15])))

    if not critical_items and not warning_items and not info_items:
        report.findings.append(Finding("Interface Errors", "OK",
            "No CRC, invalid word, FEC, or link errors found on any interface."))


def check_transceivers(outputs: dict, report: SwitchReport):
    """Check transceiver health from show interface transceiver details."""
    raw = outputs.get("show interface transceiver details", "")
    if not raw:
        report.findings.append(Finding("Transceivers", "INFO", "No transceiver data available."))
        return

    alerts = []
    current_intf = ""
    for line in raw.splitlines():
        m = re.match(r"^(fc\d+/\d+)", line, re.I)
        if m:
            current_intf = m.group(1)
        # Look for out-of-range indicators (-- or ++ or **)
        if re.search(r"[*+-]{2}", line) and current_intf:
            alerts.append(f"{current_intf}: {line.strip()}")

    if alerts:
        report.findings.append(Finding("Transceivers", "WARNING",
            f"{len(alerts)} transceiver(s) reporting out-of-range values.",
            "\n".join(alerts[:15])))
    else:
        report.findings.append(Finding("Transceivers", "OK",
            "All transceivers within normal operating range."))


def check_cpu_memory(outputs: dict, report: SwitchReport):
    """Check CPU and memory utilization."""
    raw_cpu = outputs.get("show process cpu sort", "")
    raw_mem = outputs.get("show system resources", "")

    # CPU check
    m = re.search(r"CPU utilization for five seconds:\s+(\d+)%", raw_cpu)
    if not m:
        m = re.search(r"five seconds:\s+(\d+(?:\.\d+)?)%", raw_cpu)
    if m:
        cpu = float(m.group(1))
        if cpu > 80:
            report.findings.append(Finding("CPU", "CRITICAL", f"CPU utilization is {cpu}%."))
        elif cpu > 60:
            report.findings.append(Finding("CPU", "WARNING", f"CPU utilization is {cpu}%."))
        else:
            report.findings.append(Finding("CPU", "OK", f"CPU utilization is {cpu}%."))
    else:
        report.findings.append(Finding("CPU", "INFO", "Could not determine CPU utilization."))

    # Memory check
    m_total = re.search(r"Memory usage:\s+(\d+)K total,\s+(\d+)K used,\s+(\d+)K free", raw_mem)
    if m_total:
        total = int(m_total.group(1))
        used = int(m_total.group(2))
        pct = round(used / total * 100, 1) if total else 0
        if pct > 90:
            report.findings.append(Finding("Memory", "CRITICAL", f"Memory usage at {pct}%."))
        elif pct > 75:
            report.findings.append(Finding("Memory", "WARNING", f"Memory usage at {pct}%."))
        else:
            report.findings.append(Finding("Memory", "OK", f"Memory usage at {pct}%."))
    else:
        report.findings.append(Finding("Memory", "INFO", "Could not determine memory utilization."))


def check_flogi(outputs: dict, report: SwitchReport):
    """Check FLOGI database for logged-in devices."""
    raw = outputs.get("show flogi database", "")
    if not raw:
        report.findings.append(Finding("FLOGI", "INFO", "No FLOGI data available."))
        return

    logins = [l for l in raw.splitlines() if re.match(r"\s*(fc\d+/\d+|port-channel)", l, re.I)]
    report.findings.append(Finding("FLOGI", "OK" if logins else "INFO",
        f"FLOGI database has {len(logins)} device login(s)."))


def _parse_wwpn_set(raw: str) -> set:
    """Extract all WWPNs (xx:xx:xx:xx:xx:xx:xx:xx) from text."""
    return set(re.findall(r"[0-9a-f]{2}(?::[0-9a-f]{2}){7}", raw, re.I))


def check_fcns_flogi_consistency(outputs: dict, report: SwitchReport):
    """Compare FCNS database against FLOGI database to find stale entries."""
    raw_flogi = outputs.get("show flogi database", "")
    raw_fcns = outputs.get("show fcns database local vsan 1-4093", "")

    if not raw_flogi or not raw_fcns:
        report.findings.append(Finding("FCNS/FLOGI Consistency", "INFO",
            "Insufficient data — need both FLOGI and FCNS outputs."))
        return

    flogi_wwpns = _parse_wwpn_set(raw_flogi)
    fcns_wwpns = _parse_wwpn_set(raw_fcns)

    if not flogi_wwpns and not fcns_wwpns:
        report.findings.append(Finding("FCNS/FLOGI Consistency", "INFO",
            "Both databases are empty — no logins to compare."))
        return

    # FCNS entries not in FLOGI = potentially stale
    stale = fcns_wwpns - flogi_wwpns
    # FLOGI entries not in FCNS = registration issue
    missing_reg = flogi_wwpns - fcns_wwpns

    issues = []
    if stale:
        issues.append(f"{len(stale)} FCNS entry(ies) with no matching FLOGI login (stale?)")
        for wwpn in sorted(stale)[:10]:
            issues.append(f"  stale FCNS: {wwpn}")
    if missing_reg:
        issues.append(f"{len(missing_reg)} FLOGI login(s) not registered in FCNS")
        for wwpn in sorted(missing_reg)[:10]:
            issues.append(f"  unregistered FLOGI: {wwpn}")

    if issues:
        report.findings.append(Finding("FCNS/FLOGI Consistency", "WARNING",
            f"Database mismatch: {len(stale)} stale FCNS, {len(missing_reg)} unregistered FLOGI.",
            "\n".join(issues)))
    else:
        report.findings.append(Finding("FCNS/FLOGI Consistency", "OK",
            f"FCNS and FLOGI databases are consistent ({len(flogi_wwpns)} entries)."))


def check_vsan(outputs: dict, report: SwitchReport):
    """Check VSAN states."""
    raw = outputs.get("show vsan", "")
    if not raw:
        return

    suspended = re.findall(r"VSAN\s+(\d+)\s+.*suspended", raw, re.I)
    if suspended:
        report.findings.append(Finding("VSAN", "WARNING",
            f"{len(suspended)} VSAN(s) in suspended state: {', '.join(suspended[:10])}"))
    else:
        report.findings.append(Finding("VSAN", "OK", "All VSANs active."))


def check_zoneset(outputs: dict, report: SwitchReport):
    """Check if there's an active zoneset."""
    raw = outputs.get("show zoneset active", "")
    raw_status = outputs.get("show zone status", "")

    if raw and re.search(r"zoneset name", raw, re.I):
        report.findings.append(Finding("Zoning", "OK", "Active zoneset is configured."))
    else:
        report.findings.append(Finding("Zoning", "WARNING",
            "No active zoneset found — devices may not be able to communicate."))

    # Check for zone merge failures
    if raw_status and re.search(r"merge\s+fail", raw_status, re.I):
        report.findings.append(Finding("Zoning", "CRITICAL", "Zone merge failure detected."))


def check_clock(outputs: dict, report: SwitchReport):
    """Check clock configuration (NTP recommended)."""
    raw_config = outputs.get("show running-config", "")

    if re.search(r"ntp\s+server", raw_config, re.I):
        report.findings.append(Finding("Clock / NTP", "OK", "NTP server is configured."))
    else:
        report.findings.append(Finding("Clock / NTP", "WARNING",
            "No NTP server configured — clock sync recommended for log correlation."))


def check_logging(outputs: dict, report: SwitchReport):
    """Scan recent logs for critical/error messages."""
    raw = outputs.get("show logging last 50", "")
    if not raw:
        return

    critical_lines = []
    for line in raw.splitlines():
        if re.search(r"%(ETHPORT|SYSMGR|PLATFORM|MODULE|PFMA|LC_ERROR).*(?:ERR|CRIT|FAIL)", line, re.I):
            critical_lines.append(line.strip())

    if critical_lines:
        report.findings.append(Finding("Syslog", "WARNING",
            f"{len(critical_lines)} critical/error syslog(s) in recent logs.",
            "\n".join(critical_lines[:10])))
    else:
        report.findings.append(Finding("Syslog", "OK", "No critical syslog messages in recent history."))


def check_port_channels(outputs: dict, report: SwitchReport):
    """Comprehensive port-channel health: oper status, member consistency,
    partially bundled members, speed mismatches, and member error counters."""
    raw_summary = outputs.get("show port-channel summary", "")
    raw_database = outputs.get("show port-channel database", "")
    raw_intf = outputs.get("show interface", "")
    raw_brief = outputs.get("show interface brief", "")
    raw_consistency = outputs.get("show port-channel consistency", "")

    if (not raw_summary or "No port-channel" in raw_summary) and not raw_database:
        report.findings.append(Finding("Port-Channels", "INFO", "No port-channels configured."))
        return

    critical_issues = []
    warning_issues = []
    details = []

    # ---- Parse port-channel summary for oper status ----
    pc_info = {}  # {pc_id: {"name": str, "oper": str}}
    for line in raw_summary.splitlines():
        m = re.match(r"\s*(\d+)\s+(\S+)\s+(\S+)", line)
        if m:
            pc_id, name, oper = m.group(1), m.group(2), m.group(3).lower()
            pc_info[pc_id] = {"name": name, "oper": oper}
            if "down" in oper:
                critical_issues.append(f"port-channel{pc_id} ({name}) is operationally DOWN")

    # ---- Parse port-channel database for member details ----
    # show port-channel database format:
    #   port-channel10
    #     Administrative channel mode is active
    #     Operational channel mode is active
    #     Last membership update succeeded
    #     First operational port is fc1/1
    #     2 ports in total, 2 ports up
    #       fc1/1 [up]
    #       fc1/2 [up]
    current_pc = ""
    total_ports = 0
    up_ports = 0
    member_states = []  # (intf, state)

    for line in raw_database.splitlines():
        pm = re.match(r"^(port-channel\d+)", line, re.I)
        if pm:
            # Process previous PC if it had issues
            if current_pc and total_ports > 0 and up_ports < total_ports:
                warning_issues.append(
                    f"{current_pc}: only {up_ports}/{total_ports} members bundled")
                for intf, state in member_states:
                    if state != "up":
                        warning_issues.append(f"  {intf} [{state}]")
            current_pc = pm.group(1)
            total_ports = 0
            up_ports = 0
            member_states = []
            continue

        # "2 ports in total, 2 ports up"
        count_m = re.search(r"(\d+)\s+ports?\s+in total.*?(\d+)\s+ports?\s+up", line, re.I)
        if count_m:
            total_ports = int(count_m.group(1))
            up_ports = int(count_m.group(2))
            continue

        # "  fc1/1 [up]" or "  fc1/2 [down] *"
        mem_m = re.match(r"\s+(fc\d+/\d+)\s+\[(\w+)\]", line, re.I)
        if mem_m:
            member_states.append((mem_m.group(1), mem_m.group(2).lower()))

    # Process last PC
    if current_pc and total_ports > 0 and up_ports < total_ports:
        warning_issues.append(
            f"{current_pc}: only {up_ports}/{total_ports} members bundled")
        for intf, state in member_states:
            if state != "up":
                warning_issues.append(f"  {intf} [{state}]")

    # ---- Check for speed mismatches in members from show interface brief ----
    # Build {intf: speed} map
    intf_speeds = {}
    for line in raw_brief.splitlines():
        sm = re.match(r"\s*(fc\d+/\d+)\s+\S+\s+\S+\s+(\S+)\s+(\d+[GMTK]?)", line, re.I)
        if sm:
            intf_speeds[sm.group(1).lower()] = sm.group(3)

    # Re-parse database to check per-PC speed consistency
    current_pc = ""
    pc_members = []
    for line in raw_database.splitlines():
        pm = re.match(r"^(port-channel\d+)", line, re.I)
        if pm:
            if current_pc and pc_members:
                speeds = {intf_speeds.get(m.lower(), "?") for m in pc_members}
                if len(speeds) > 1:
                    speed_detail = ", ".join(f"{m}={intf_speeds.get(m.lower(),'?')}" for m in pc_members)
                    warning_issues.append(f"{current_pc}: speed mismatch across members: {speed_detail}")
            current_pc = pm.group(1)
            pc_members = []
            continue
        mem_m = re.match(r"\s+(fc\d+/\d+)\s+\[\w+\]", line, re.I)
        if mem_m:
            pc_members.append(mem_m.group(1))
    # Process last
    if current_pc and pc_members:
        speeds = {intf_speeds.get(m.lower(), "?") for m in pc_members}
        if len(speeds) > 1:
            speed_detail = ", ".join(f"{m}={intf_speeds.get(m.lower(),'?')}" for m in pc_members)
            warning_issues.append(f"{current_pc}: speed mismatch across members: {speed_detail}")

    # ---- Check port-channel consistency output ----
    if raw_consistency:
        for line in raw_consistency.splitlines():
            if re.search(r"(fail|mismatch|inconsisten)", line, re.I):
                warning_issues.append(f"Consistency: {line.strip()}")

    # ---- Report findings ----
    total_pcs = len(pc_info) if pc_info else "?"
    if critical_issues:
        report.findings.append(Finding("Port-Channels", "CRITICAL",
            f"{len(critical_issues)} port-channel(s) DOWN.",
            "\n".join(critical_issues)))
    if warning_issues:
        report.findings.append(Finding("Port-Channels", "WARNING",
            f"{len(warning_issues)} port-channel issue(s) (partial bundle / speed mismatch).",
            "\n".join(warning_issues[:20])))
    if not critical_issues and not warning_issues:
        report.findings.append(Finding("Port-Channels", "OK",
            f"All {total_pcs} port-channel(s) fully operational — all members bundled."))


def check_device_alias(outputs: dict, report: SwitchReport):
    """Check device-alias configuration, distribution mode, and pending changes."""
    raw_status = outputs.get("show device-alias status", "")
    raw_db = outputs.get("show device-alias database", "")
    raw_flogi = outputs.get("show flogi database", "")

    if not raw_status:
        report.findings.append(Finding("Device-Alias", "INFO", "No device-alias status data available."))
        return

    details = []

    # Check distribution mode (enhanced vs basic)
    if re.search(r"Fabric Distribution:\s*Enabled", raw_status, re.I):
        details.append("Distribution mode: Enabled (good)")
    elif re.search(r"Fabric Distribution:\s*Disabled", raw_status, re.I):
        report.findings.append(Finding("Device-Alias", "WARNING",
            "Device-alias fabric distribution is DISABLED.",
            "Enable with: device-alias distribute"))
    else:
        details.append("Distribution mode: unknown")

    # Check device-alias mode (enhanced vs basic)
    if re.search(r"Device Alias Mode:\s*Enhanced", raw_status, re.I):
        details.append("Mode: Enhanced")
    elif re.search(r"Device Alias Mode:\s*Basic", raw_status, re.I):
        details.append("Mode: Basic")

    # Check for pending changes (locked session)
    if re.search(r"Locked\s+By", raw_status, re.I):
        m = re.search(r"Locked\s+By:\s*(\S+)", raw_status, re.I)
        locked_by = m.group(1) if m else "unknown"
        report.findings.append(Finding("Device-Alias", "WARNING",
            f"Device-alias database has a pending session locked by {locked_by}.",
            "Uncommitted changes can block other administrators.\n"
            "Commit with: device-alias commit\n"
            "Abort with: device-alias abort"))

    # Count aliases and compare to FLOGI logins
    alias_count = 0
    alias_wwpns = set()
    if raw_db:
        for line in raw_db.splitlines():
            m = re.match(r"\s*(\S+)\s+([0-9a-f]{2}(?::[0-9a-f]{2}){7})", line, re.I)
            if m:
                alias_count += 1
                alias_wwpns.add(m.group(2).lower())

    details.append(f"Total aliases defined: {alias_count}")

    # Find FLOGI devices without aliases
    if raw_flogi and alias_wwpns:
        flogi_wwpns = _parse_wwpn_set(raw_flogi)
        no_alias = flogi_wwpns - alias_wwpns
        if no_alias:
            detail_lines = [f"{len(no_alias)} logged-in device(s) have no device-alias:"]
            for wwpn in sorted(no_alias)[:10]:
                detail_lines.append(f"  {wwpn}")
            if len(no_alias) > 10:
                detail_lines.append(f"  ... and {len(no_alias) - 10} more")
            report.findings.append(Finding("Device-Alias", "INFO",
                f"{len(no_alias)} FLOGI device(s) without a device-alias.",
                "\n".join(detail_lines)))
        else:
            details.append("All logged-in devices have aliases")

    sev = "OK" if not any(f.module == "Device-Alias" and f.severity in ("WARNING", "CRITICAL")
                          for f in report.findings) else "INFO"
    report.findings.append(Finding("Device-Alias", sev,
        f"Device-alias database: {alias_count} entries.",
        "\n".join(details)))


def check_smart_zoning(outputs: dict, report: SwitchReport):
    """Check whether smart zoning is enabled per VSAN."""
    raw = outputs.get("show zone status vsan 1-4093", "")
    if not raw:
        raw = outputs.get("show zone status", "")
    if not raw:
        report.findings.append(Finding("Smart Zoning", "INFO", "No zone status data available."))
        return

    # Parse per-VSAN smart zoning status
    current_vsan = None
    smart_enabled = []
    smart_disabled = []

    for line in raw.splitlines():
        vm = re.search(r"VSAN:\s*(\d+)", line, re.I)
        if vm:
            current_vsan = vm.group(1)
        if current_vsan:
            if re.search(r"smart.?zoning\s*:\s*Enabled", line, re.I):
                smart_enabled.append(current_vsan)
                current_vsan = None
            elif re.search(r"smart.?zoning\s*:\s*Disabled", line, re.I):
                smart_disabled.append(current_vsan)
                current_vsan = None

    details = []
    if smart_enabled:
        details.append(f"Smart zoning ENABLED on VSANs: {', '.join(smart_enabled[:20])}")
    if smart_disabled:
        details.append(f"Smart zoning DISABLED on VSANs: {', '.join(smart_disabled[:20])}")

    if smart_disabled:
        report.findings.append(Finding("Smart Zoning", "WARNING",
            f"Smart zoning disabled on {len(smart_disabled)} VSAN(s) — "
            "consider enabling to reduce TCAM usage.",
            "\n".join(details)))
    elif smart_enabled:
        report.findings.append(Finding("Smart Zoning", "OK",
            f"Smart zoning enabled on all {len(smart_enabled)} active VSAN(s)."))
    else:
        report.findings.append(Finding("Smart Zoning", "INFO",
            "Could not determine smart zoning status from zone output."))


def check_acl_tcam(outputs: dict, report: SwitchReport):
    """Check ACL / TCAM resource utilization on linecards."""
    raw = outputs.get("show hardware internal fcmac acl-sobp resource-usage", "")
    if not raw:
        report.findings.append(Finding("ACL TCAM", "INFO",
            "No TCAM resource data available (command may not be supported on this platform)."))
        return

    issues = []
    details = []
    current_module = ""

    for line in raw.splitlines():
        mm = re.search(r"Module\s+(\d+)", line, re.I)
        if mm:
            current_module = mm.group(1)

        # Match percentage patterns — various MDS output formats
        # e.g., "sobp entries used: 1500 out of 2048 (73%)"
        # e.g., "Used:  1500/2048 (73.24%)"
        pct_match = re.search(r"(\d+(?:\.\d+)?)\s*%", line)
        if pct_match:
            pct = float(pct_match.group(1))
            label = line.strip()[:80]
            mod_label = f"Module {current_module}" if current_module else "unknown"

            if pct >= 90:
                issues.append(f"CRITICAL — {mod_label}: {label}")
            elif pct >= 75:
                issues.append(f"WARNING  — {mod_label}: {label}")
            elif pct >= 50:
                details.append(f"{mod_label}: {label}")

        # Also catch "sobp" / "acl" entries with used/total format
        usage_match = re.search(r"(\d+)\s+(?:out of|/)\s+(\d+)", line, re.I)
        if usage_match and not pct_match:
            used = int(usage_match.group(1))
            total = int(usage_match.group(2))
            if total > 0:
                pct = round(used / total * 100, 1)
                mod_label = f"Module {current_module}" if current_module else "unknown"
                label = f"{line.strip()[:60]} ({pct}%)"
                if pct >= 90:
                    issues.append(f"CRITICAL — {mod_label}: {label}")
                elif pct >= 75:
                    issues.append(f"WARNING  — {mod_label}: {label}")
                elif pct >= 50:
                    details.append(f"{mod_label}: {label}")

    if issues:
        crit_count = sum(1 for i in issues if i.startswith("CRITICAL"))
        sev = "CRITICAL" if crit_count > 0 else "WARNING"
        report.findings.append(Finding("ACL TCAM", sev,
            f"{len(issues)} TCAM resource(s) at high utilization.",
            "\n".join(issues[:15])))
    else:
        report.findings.append(Finding("ACL TCAM", "OK",
            "TCAM/ACL resource utilization is within normal limits.",
            "\n".join(details[:5]) if details else ""))


# ─────────────────────────────────────────────
# Fabric zoning comparison (cross-switch)
# ─────────────────────────────────────────────

def _parse_zoneset_zones(raw: str) -> dict:
    """Parse 'show zoneset active' into {zoneset: {zone: set(members)}}."""
    result = {}
    current_zoneset = None
    current_zone = None

    for line in raw.splitlines():
        zs = re.match(r"\s*zoneset\s+name\s+(\S+)\s+vsan\s+(\d+)", line, re.I)
        if zs:
            current_zoneset = f"{zs.group(1)}(vsan{zs.group(2)})"
            result[current_zoneset] = {}
            current_zone = None
            continue
        zm = re.match(r"\s*zone\s+name\s+(\S+)\s+vsan\s+(\d+)", line, re.I)
        if zm and current_zoneset:
            current_zone = zm.group(1)
            result[current_zoneset][current_zone] = set()
            continue
        # Members: pwwn, device-alias, fcid, etc.
        mm = re.match(r"\s+(pwwn|device-alias|fcid|interface)\s+(\S+)", line, re.I)
        if mm and current_zoneset and current_zone:
            member_type = mm.group(1).lower()
            member_val = mm.group(2).lower()
            result[current_zoneset][current_zone].add(f"{member_type}:{member_val}")

    return result


def compare_fabrics(outputs_a: dict, outputs_b: dict,
                    label_a: str = "Fabric-A", label_b: str = "Fabric-B") -> list:
    """Compare zoning between two fabrics and return a list of Findings."""
    findings = []

    raw_a = outputs_a.get("show zoneset active vsan 1-4093", "") or outputs_a.get("show zoneset active", "")
    raw_b = outputs_b.get("show zoneset active vsan 1-4093", "") or outputs_b.get("show zoneset active", "")

    zs_a = _parse_zoneset_zones(raw_a)
    zs_b = _parse_zoneset_zones(raw_b)

    all_zonesets = set(zs_a.keys()) | set(zs_b.keys())

    if not all_zonesets:
        findings.append(Finding("Fabric Zoning Compare", "INFO",
            "No active zonesets found on either fabric to compare."))
        return findings

    details = []

    # ---- Zoneset-level comparison ----
    only_a = set(zs_a.keys()) - set(zs_b.keys())
    only_b = set(zs_b.keys()) - set(zs_a.keys())

    if only_a:
        details.append(f"Zonesets only in {label_a}: {', '.join(sorted(only_a))}")
    if only_b:
        details.append(f"Zonesets only in {label_b}: {', '.join(sorted(only_b))}")

    # ---- Zone-level comparison within matching zonesets ----
    common_zonesets = set(zs_a.keys()) & set(zs_b.keys())
    zone_diffs = []
    member_diffs = []

    for zs_name in sorted(common_zonesets):
        zones_a = set(zs_a[zs_name].keys())
        zones_b = set(zs_b[zs_name].keys())

        z_only_a = zones_a - zones_b
        z_only_b = zones_b - zones_a

        if z_only_a:
            zone_diffs.append(f"  {zs_name}: zones only in {label_a}: {', '.join(sorted(z_only_a)[:5])}")
        if z_only_b:
            zone_diffs.append(f"  {zs_name}: zones only in {label_b}: {', '.join(sorted(z_only_b)[:5])}")

        # ---- Member-level comparison within matching zones ----
        common_zones = zones_a & zones_b
        for zone_name in sorted(common_zones):
            members_a = zs_a[zs_name][zone_name]
            members_b = zs_b[zs_name][zone_name]
            m_only_a = members_a - members_b
            m_only_b = members_b - members_a
            if m_only_a or m_only_b:
                diff_detail = f"  {zs_name} > {zone_name}:"
                if m_only_a:
                    diff_detail += f" only-{label_a}=[{', '.join(sorted(m_only_a)[:3])}]"
                if m_only_b:
                    diff_detail += f" only-{label_b}=[{', '.join(sorted(m_only_b)[:3])}]"
                member_diffs.append(diff_detail)

    if zone_diffs:
        details.append("Zone differences:")
        details.extend(zone_diffs[:15])
    if member_diffs:
        details.append("Member differences:")
        details.extend(member_diffs[:15])

    total_diffs = len(only_a) + len(only_b) + len(zone_diffs) + len(member_diffs)

    if total_diffs > 0:
        findings.append(Finding("Fabric Zoning Compare",
            "WARNING" if total_diffs < 10 else "CRITICAL",
            f"{total_diffs} zoning difference(s) between {label_a} and {label_b}.",
            "\n".join(details)))
    else:
        findings.append(Finding("Fabric Zoning Compare", "OK",
            f"Zoning is consistent between {label_a} and {label_b} "
            f"({len(common_zonesets)} zoneset(s) compared)."))

    # ---- Also compare device-alias databases ----
    da_a = _parse_wwpn_set(outputs_a.get("show device-alias database", ""))
    da_b = _parse_wwpn_set(outputs_b.get("show device-alias database", ""))
    if da_a or da_b:
        da_only_a = da_a - da_b
        da_only_b = da_b - da_a
        if da_only_a or da_only_b:
            da_details = []
            if da_only_a:
                da_details.append(f"Aliases only in {label_a}: {len(da_only_a)}")
            if da_only_b:
                da_details.append(f"Aliases only in {label_b}: {len(da_only_b)}")
            findings.append(Finding("Fabric Device-Alias Compare", "WARNING",
                f"Device-alias mismatch between fabrics.",
                "\n".join(da_details)))
        else:
            findings.append(Finding("Fabric Device-Alias Compare", "OK",
                f"Device-alias databases match ({len(da_a)} entries)."))

    return findings


# ─────────────────────────────────────────────
# ISL & linecard load balancing
# ─────────────────────────────────────────────

def _parse_interface_rates(raw_brief: str, raw_full: str) -> list:
    """Parse interface rates and return list of dicts with rate info.
    Returns: [{intf, module, speed_gbps, in_rate_bps, out_rate_bps, mode, vsan}]
    """
    interfaces = []

    # Parse show interface brief for port type (TE/TF/E/F/ISL) and speed
    brief_info = {}  # {intf: {speed, mode, vsan, oper_status}}
    for line in raw_brief.splitlines():
        # fc1/1  1  32  E  auto  TE  trunk  up  --
        m = re.match(
            r"\s*(fc\d+/\d+)\s+(\d+)\s+(\d+[GMTK]?)\s+(\S+)\s+\S+\s+(\S+)\s+\S+\s+(\S+)",
            line, re.I
        )
        if m:
            intf = m.group(1)
            vsan = m.group(2)
            speed = m.group(3)
            port_mode = m.group(4)  # E, F, TE, TF, etc.
            trunk_mode = m.group(5)
            oper = m.group(6).lower()
            brief_info[intf.lower()] = {
                "speed": speed, "mode": port_mode, "trunk": trunk_mode,
                "vsan": vsan, "oper": oper
            }

    # Parse show interface for input/output rates
    current_intf = ""
    current_in_bps = 0
    current_out_bps = 0

    for line in raw_full.splitlines():
        im = re.match(r"^(fc\d+/\d+)\s+is\s+(up|down)", line, re.I)
        if im:
            # Save previous
            if current_intf:
                info = brief_info.get(current_intf.lower(), {})
                module = re.match(r"fc(\d+)/", current_intf)
                mod_num = module.group(1) if module else "?"
                interfaces.append({
                    "intf": current_intf,
                    "module": mod_num,
                    "speed": info.get("speed", "?"),
                    "mode": info.get("mode", "?"),
                    "vsan": info.get("vsan", "?"),
                    "in_rate_bps": current_in_bps,
                    "out_rate_bps": current_out_bps,
                })
            current_intf = im.group(1)
            current_in_bps = 0
            current_out_bps = 0
            continue

        # "  30 seconds input rate 1234567 bits/sec, 1234 bytes/sec, 100 frames/sec"
        rate_in = re.search(r"input rate\s+(\d+)\s+bits", line, re.I)
        if rate_in:
            current_in_bps = int(rate_in.group(1))
        rate_out = re.search(r"output rate\s+(\d+)\s+bits", line, re.I)
        if rate_out:
            current_out_bps = int(rate_out.group(1))

    # Save last interface
    if current_intf:
        info = brief_info.get(current_intf.lower(), {})
        module = re.match(r"fc(\d+)/", current_intf)
        mod_num = module.group(1) if module else "?"
        interfaces.append({
            "intf": current_intf,
            "module": mod_num,
            "speed": info.get("speed", "?"),
            "mode": info.get("mode", "?"),
            "vsan": info.get("vsan", "?"),
            "in_rate_bps": current_in_bps,
            "out_rate_bps": current_out_bps,
        })

    return interfaces


def _speed_to_bps(speed_str: str) -> int:
    """Convert speed string like '32G' to bits per second."""
    m = re.match(r"(\d+)\s*([GMTK])?", speed_str, re.I)
    if not m:
        return 0
    val = int(m.group(1))
    suffix = (m.group(2) or "G").upper()
    multiplier = {"K": 1_000, "M": 1_000_000, "G": 1_000_000_000, "T": 1_000_000_000_000}
    return val * multiplier.get(suffix, 1_000_000_000)


def check_isl_load_balance(outputs: dict, report: SwitchReport):
    """Analyze traffic distribution across ISL/TE/TF ports and port-channel
    members. Detect skew that indicates poor load balancing."""
    raw_brief = outputs.get("show interface brief", "")
    raw_full = outputs.get("show interface", "")
    raw_pc_db = outputs.get("show port-channel database", "")

    interfaces = _parse_interface_rates(raw_brief, raw_full)
    if not interfaces:
        report.findings.append(Finding("ISL Load Balance", "INFO",
            "No interface rate data available for load balance analysis."))
        return

    # ---- ISL / trunk port identification ----
    isl_ports = [i for i in interfaces
                 if i["mode"].upper() in ("E", "TE", "TF", "TP")]

    if not isl_ports:
        report.findings.append(Finding("ISL Load Balance", "INFO",
            "No ISL/E/TE/TF ports found — skipping ISL load balance check."))
    else:
        # ---- Per-port-channel member balance ----
        # Parse port-channel membership
        pc_members = {}  # {pc_name: [intf, ...]}
        current_pc = ""
        for line in raw_pc_db.splitlines():
            pm = re.match(r"^(port-channel\d+)", line, re.I)
            if pm:
                current_pc = pm.group(1)
                pc_members[current_pc] = []
                continue
            mem_m = re.match(r"\s+(fc\d+/\d+)\s+\[(\w+)\]", line, re.I)
            if mem_m and current_pc:
                pc_members[current_pc].append(mem_m.group(1))

        pc_balance_issues = []
        pc_balance_ok = []

        intf_rate_map = {i["intf"].lower(): i for i in interfaces}

        for pc_name, members in sorted(pc_members.items()):
            if len(members) < 2:
                continue

            member_rates = []
            for m_intf in members:
                info = intf_rate_map.get(m_intf.lower())
                if info:
                    total = info["in_rate_bps"] + info["out_rate_bps"]
                    member_rates.append((m_intf, total))

            if not member_rates:
                continue

            rates = [r for _, r in member_rates]
            total_rate = sum(rates)
            if total_rate == 0:
                pc_balance_ok.append(f"{pc_name}: idle (0 bps across {len(members)} members)")
                continue

            avg_rate = total_rate / len(rates)
            max_rate = max(rates)
            min_rate = min(rates)

            # Skew metric: if any member carries >2x the average = imbalanced
            if avg_rate > 0:
                skew = max_rate / avg_rate
            else:
                skew = 1.0

            # Also check if one member handles >60% of total traffic
            max_pct = (max_rate / total_rate * 100) if total_rate else 0

            if skew > 3.0 or max_pct > 70:
                sev = "WARNING"
                rate_detail = ", ".join(
                    f"{intf}={rate/1_000_000:.1f}Mbps" for intf, rate in member_rates)
                pc_balance_issues.append(
                    f"{pc_name}: SKEWED — max member carries {max_pct:.0f}% "
                    f"(skew={skew:.1f}x) [{rate_detail}]")
            else:
                pc_balance_ok.append(
                    f"{pc_name}: balanced ({len(members)} members, "
                    f"total={total_rate/1_000_000:.1f}Mbps, skew={skew:.1f}x)")

        if pc_balance_issues:
            report.findings.append(Finding("ISL Load Balance", "WARNING",
                f"{len(pc_balance_issues)} port-channel(s) with uneven traffic distribution.",
                "\n".join(pc_balance_issues[:10])))
        elif pc_balance_ok:
            report.findings.append(Finding("ISL Load Balance", "OK",
                f"Traffic evenly distributed across {len(pc_balance_ok)} port-channel(s).",
                "\n".join(pc_balance_ok[:5])))

    # ---- Per-ISL utilization check (any ISL >80% capacity) ----
    overloaded_isls = []
    for port in isl_ports:
        speed_bps = _speed_to_bps(port["speed"])
        if speed_bps == 0:
            continue
        max_rate = max(port["in_rate_bps"], port["out_rate_bps"])
        util_pct = max_rate / speed_bps * 100
        if util_pct > 80:
            overloaded_isls.append(
                f"{port['intf']}: {util_pct:.0f}% utilization "
                f"(in={port['in_rate_bps']/1_000_000:.1f}M, "
                f"out={port['out_rate_bps']/1_000_000:.1f}M, "
                f"speed={port['speed']})")

    if overloaded_isls:
        report.findings.append(Finding("ISL Utilization", "WARNING",
            f"{len(overloaded_isls)} ISL(s) above 80% utilization.",
            "\n".join(overloaded_isls[:10])))
    elif isl_ports:
        report.findings.append(Finding("ISL Utilization", "OK",
            f"All {len(isl_ports)} ISL/trunk port(s) below 80% utilization."))


def check_linecard_balance(outputs: dict, report: SwitchReport):
    """Analyze traffic distribution across linecards (modules) to detect
    hot linecards versus idle ones."""
    raw_brief = outputs.get("show interface brief", "")
    raw_full = outputs.get("show interface", "")

    interfaces = _parse_interface_rates(raw_brief, raw_full)
    if not interfaces:
        report.findings.append(Finding("Linecard Balance", "INFO",
            "No interface data for linecard analysis."))
        return

    # Aggregate rates per module
    module_rates = {}  # {module: {total_in, total_out, port_count, active_ports}}
    for intf in interfaces:
        mod = intf["module"]
        if mod == "?":
            continue
        entry = module_rates.setdefault(mod, {
            "total_in": 0, "total_out": 0, "port_count": 0, "active_ports": 0
        })
        entry["total_in"] += intf["in_rate_bps"]
        entry["total_out"] += intf["out_rate_bps"]
        entry["port_count"] += 1
        if intf["in_rate_bps"] + intf["out_rate_bps"] > 0:
            entry["active_ports"] += 1

    if len(module_rates) < 2:
        report.findings.append(Finding("Linecard Balance", "INFO",
            "Only one linecard with traffic — skipping balance analysis."))
        return

    # Calculate total and per-module percentages
    grand_total = sum(m["total_in"] + m["total_out"] for m in module_rates.values())
    if grand_total == 0:
        report.findings.append(Finding("Linecard Balance", "OK",
            "All linecards idle — no traffic to analyze for balance."))
        return

    details = []
    max_pct = 0
    for mod_id in sorted(module_rates.keys()):
        m = module_rates[mod_id]
        mod_total = m["total_in"] + m["total_out"]
        pct = mod_total / grand_total * 100
        max_pct = max(max_pct, pct)
        details.append(
            f"Module {mod_id}: {mod_total/1_000_000:.1f} Mbps "
            f"({pct:.0f}% of total, {m['active_ports']}/{m['port_count']} active ports)")

    # If any single linecard handles >65% of all traffic with >2 cards, flag it
    if max_pct > 65 and len(module_rates) > 1:
        report.findings.append(Finding("Linecard Balance", "WARNING",
            f"Uneven traffic distribution — one linecard carries {max_pct:.0f}% of total.",
            "\n".join(details)))
    else:
        report.findings.append(Finding("Linecard Balance", "OK",
            f"Traffic distributed across {len(module_rates)} linecards "
            f"(max {max_pct:.0f}% on any single module).",
            "\n".join(details[:8])))


# ─────────────────────────────────────────────
# Report rendering
# ─────────────────────────────────────────────

SEPARATOR = "=" * 72

def render_text_report(report: SwitchReport) -> str:
    """Render a human-readable text report."""
    lines = []
    lines.append("")
    lines.append(SEPARATOR)
    lines.append(f"  CISCO MDS HEALTH CHECK REPORT")
    lines.append(SEPARATOR)
    lines.append(f"  Switch      : {report.hostname} ({report.ip})")
    lines.append(f"  Model       : {report.model}")
    lines.append(f"  NX-OS       : {report.nxos_version}")
    lines.append(f"  Uptime      : {report.uptime}")
    lines.append(f"  Checked at  : {report.timestamp}")
    lines.append(f"  Status      : {Colors.severity(report.overall_status)}")
    lines.append(SEPARATOR)
    lines.append("")

    # Group by severity
    for sev in ("CRITICAL", "WARNING", "INFO", "OK"):
        items = [f for f in report.findings if f.severity == sev]
        if not items:
            continue
        lines.append(f"  [{Colors.severity(sev)}]  ({len(items)} finding{'s' if len(items)!=1 else ''})")
        lines.append(f"  {'-' * 60}")
        for f in items:
            lines.append(f"    {f.module}: {f.message}")
            if f.detail:
                for dl in f.detail.splitlines()[:5]:
                    lines.append(f"      {Colors.DIM}{dl}{Colors.RESET}")
        lines.append("")

    lines.append(SEPARATOR)
    lines.append(f"  Summary: {report.critical_count} critical, {report.warning_count} warning, "
                 f"{len(report.findings)} total checks")
    lines.append(SEPARATOR)
    lines.append("")
    return "\n".join(lines)


def render_json_report(report: SwitchReport) -> str:
    """Render a JSON report."""
    data = {
        "hostname": report.hostname,
        "ip": report.ip,
        "model": report.model,
        "nxos_version": report.nxos_version,
        "uptime": report.uptime,
        "timestamp": report.timestamp,
        "overall_status": report.overall_status,
        "critical_count": report.critical_count,
        "warning_count": report.warning_count,
        "findings": [asdict(f) for f in report.findings],
    }
    return json.dumps(data, indent=2)


def render_xlsx_report(reports: list, filepath: str):
    """Export one or more SwitchReports to a formatted Excel workbook.
    Sheets: Summary, then one detail sheet per switch."""
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl is required for Excel export. Install with: pip install openpyxl")
        return

    wb = Workbook()

    # ---- Styles ----
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    sev_fills = {
        "CRITICAL": PatternFill("solid", fgColor="FF4444"),
        "WARNING":  PatternFill("solid", fgColor="FFC000"),
        "INFO":     PatternFill("solid", fgColor="5B9BD5"),
        "OK":       PatternFill("solid", fgColor="70AD47"),
    }
    sev_fonts = {
        "CRITICAL": Font(bold=True, color="FFFFFF"),
        "WARNING":  Font(bold=True, color="000000"),
        "INFO":     Font(color="FFFFFF"),
        "OK":       Font(color="FFFFFF"),
    }
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")

    def style_header_row(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ═══════════════════════════════════════════
    # Summary sheet
    # ═══════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = ["Switch", "IP", "Model", "NX-OS", "Uptime", "Status",
                   "Critical", "Warning", "Total Checks", "Timestamp"]
    ws_sum.append(sum_headers)
    style_header_row(ws_sum, len(sum_headers))

    for r in reports:
        row = [
            r.hostname, r.ip, r.model, r.nxos_version, r.uptime,
            r.overall_status, r.critical_count, r.warning_count,
            len(r.findings), r.timestamp,
        ]
        ws_sum.append(row)
        row_num = ws_sum.max_row
        # Color the status cell
        status_cell = ws_sum.cell(row=row_num, column=6)
        sev = r.overall_status
        if sev in sev_fills:
            status_cell.fill = sev_fills.get(sev, PatternFill())
            status_cell.font = sev_fonts.get(sev, Font())
        # Border all cells
        for col in range(1, len(sum_headers) + 1):
            ws_sum.cell(row=row_num, column=col).border = thin_border

    # Auto-width
    for col in range(1, len(sum_headers) + 1):
        max_len = max(len(str(ws_sum.cell(row=r, column=col).value or ""))
                      for r in range(1, ws_sum.max_row + 1))
        ws_sum.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # ═══════════════════════════════════════════
    # Per-switch detail sheets
    # ═══════════════════════════════════════════
    for report in reports:
        sheet_name = (report.hostname or report.ip or "Switch")[:31]
        # Ensure unique name
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + f"_{len(existing)}"

        ws = wb.create_sheet(title=sheet_name)

        # Info header rows
        info_rows = [
            ("Switch:", report.hostname),
            ("IP:", report.ip),
            ("Model:", report.model),
            ("NX-OS:", report.nxos_version),
            ("Uptime:", report.uptime),
            ("Status:", report.overall_status),
            ("Checked:", report.timestamp),
            ("", ""),
        ]
        for label, val in info_rows:
            ws.append([label, val])
            if label == "Status:":
                cell = ws.cell(row=ws.max_row, column=2)
                sev = report.overall_status
                if sev in sev_fills:
                    cell.fill = sev_fills[sev]
                    cell.font = sev_fonts[sev]

        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=6, column=1).font = Font(bold=True)

        # Findings table
        detail_headers = ["Severity", "Module", "Message", "Detail"]
        ws.append(detail_headers)
        header_row = ws.max_row
        for col in range(1, len(detail_headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Sort findings: CRITICAL first, then WARNING, INFO, OK
        sev_order = {"CRITICAL": 0, "WARNING": 1, "INFO": 2, "OK": 3}
        sorted_findings = sorted(report.findings, key=lambda f: sev_order.get(f.severity, 9))

        for f in sorted_findings:
            ws.append([f.severity, f.module, f.message, f.detail])
            row_num = ws.max_row
            # Severity cell coloring
            sev_cell = ws.cell(row=row_num, column=1)
            if f.severity in sev_fills:
                sev_cell.fill = sev_fills[f.severity]
                sev_cell.font = sev_fonts[f.severity]
            # Border and wrap
            for col in range(1, len(detail_headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = wrap_align

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 80

    # Save
    wb.save(filepath)
    print(f"  Excel report saved: {filepath}")


# ─────────────────────────────────────────────
# Main orchestrator
# ─────────────────────────────────────────────

ALL_CHECKS = [
    check_version,
    check_uptime,
    check_modules,
    check_environment,
    check_cpu_memory,
    check_interfaces,
    check_interface_errors,
    check_transceivers,
    check_flogi,
    check_fcns_flogi_consistency,
    check_vsan,
    check_zoneset,
    check_smart_zoning,
    check_acl_tcam,
    check_port_channels,
    check_isl_load_balance,
    check_linecard_balance,
    check_device_alias,
    check_clock,
    check_logging,
]


def health_check(ip: str, username: str, password: str, json_output: bool = False) -> Optional[SwitchReport]:
    """Run a full health check on one MDS switch."""
    conn = connect_switch(ip, username, password)
    if not conn:
        return None

    try:
        print(f"  Collecting data ({len(HEALTH_COMMANDS)} commands) ...")
        outputs = collect_outputs(conn, HEALTH_COMMANDS)
    finally:
        conn.disconnect()

    report = SwitchReport(ip=ip, timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    print(f"  Analyzing ...")
    for check_fn in ALL_CHECKS:
        try:
            check_fn(outputs, report)
        except Exception as e:
            report.findings.append(Finding(check_fn.__name__, "INFO", f"Check error: {e}"))

    if json_output:
        print(render_json_report(report))
    else:
        print(render_text_report(report))

    return report


def main():
    parser = argparse.ArgumentParser(
        description="Cisco MDS 9000 Health Check Script",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s -t 10.1.1.1
  %(prog)s -t 10.1.1.1 -u admin
  %(prog)s -f switches.txt --json
  %(prog)s -f switches.txt -o report.json --json
  %(prog)s --fabric-a 10.1.1.1 --fabric-b 10.2.2.1 -u admin   # compare fabrics
        """,
    )
    parser.add_argument("-t", "--target", help="Switch IP address or hostname")
    parser.add_argument("-f", "--file", help="File with one switch IP per line")
    parser.add_argument("-u", "--username", help="SSH username (will prompt if omitted)")
    parser.add_argument("-p", "--password", help="SSH password (will prompt if omitted)")
    parser.add_argument("-o", "--output", help="Save report to file")
    parser.add_argument("--json", action="store_true", help="Output as JSON")
    parser.add_argument("--xlsx", help="Export report to Excel (.xlsx) file")
    parser.add_argument("--fabric-a", dest="fabric_a", help="Fabric A switch IP (for zoning comparison)")
    parser.add_argument("--fabric-b", dest="fabric_b", help="Fabric B switch IP (for zoning comparison)")
    args = parser.parse_args()

    # ---- Fabric comparison mode ----
    if args.fabric_a and args.fabric_b:
        username = args.username or input("Username: ").strip()
        password = args.password or getpass.getpass("Password: ")

        print(f"\n{'=' * 72}")
        print(f"  Cisco MDS Fabric Zoning Comparison")
        print(f"  Fabric A: {args.fabric_a}    Fabric B: {args.fabric_b}")
        print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'=' * 72}\n")

        conn_a = connect_switch(args.fabric_a, username, password)
        conn_b = connect_switch(args.fabric_b, username, password)
        if not conn_a or not conn_b:
            print(f"  {Colors.RED}Could not connect to both switches. Aborting comparison.{Colors.RESET}")
            sys.exit(1)

        try:
            print(f"  Collecting from Fabric A ({args.fabric_a}) ...")
            out_a = collect_outputs(conn_a, HEALTH_COMMANDS)
        finally:
            conn_a.disconnect()
        try:
            print(f"  Collecting from Fabric B ({args.fabric_b}) ...")
            out_b = collect_outputs(conn_b, HEALTH_COMMANDS)
        finally:
            conn_b.disconnect()

        print(f"  Comparing zoning and device-alias databases ...\n")
        findings = compare_fabrics(out_a, out_b,
                                   label_a=f"FabA({args.fabric_a})",
                                   label_b=f"FabB({args.fabric_b})")

        compare_report = SwitchReport(
            hostname="Fabric Comparison",
            ip=f"{args.fabric_a} vs {args.fabric_b}",
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            findings=findings,
        )
        if args.json:
            print(render_json_report(compare_report))
        else:
            print(render_text_report(compare_report))

        if args.output:
            with open(args.output, "w") as fh:
                if args.json:
                    json.dump(asdict(compare_report), fh, indent=2, default=str)
                else:
                    fh.write(render_text_report(compare_report))
            print(f"\nComparison report saved to: {args.output}")

        if args.xlsx:
            xlsx_path = args.xlsx if args.xlsx.endswith(".xlsx") else args.xlsx + ".xlsx"
            render_xlsx_report([compare_report], xlsx_path)

        # Also run individual health checks on both
        print(f"\n  Running individual health checks on both fabrics ...\n")
        for fab_ip in (args.fabric_a, args.fabric_b):
            health_check(fab_ip, username, password, json_output=args.json)

        sys.exit(0)

    elif args.fabric_a or args.fabric_b:
        print("ERROR: --fabric-a and --fabric-b must both be specified for comparison.")
        sys.exit(1)

    # Gather targets
    targets = []
    if args.target:
        targets.append(args.target)
    elif args.file:
        if not os.path.isfile(args.file):
            print(f"ERROR: File not found: {args.file}")
            sys.exit(1)
        with open(args.file) as fh:
            targets = [l.strip() for l in fh if l.strip() and not l.startswith("#")]
    else:
        ip = input("Enter switch IP or hostname: ").strip()
        if not ip:
            print("No target specified.")
            sys.exit(1)
        targets.append(ip)

    # Credentials
    username = args.username or input("Username: ").strip()
    password = args.password or getpass.getpass("Password: ")

    print(f"\n{'=' * 72}")
    print(f"  Cisco MDS Health Check — {len(targets)} switch(es)")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 72}\n")

    reports = []
    for ip in targets:
        report = health_check(ip, username, password, json_output=args.json)
        if report:
            reports.append(report)

    # Save to file if requested
    if args.output and reports:
        with open(args.output, "w") as fh:
            if args.json:
                json.dump([asdict(r) for r in reports], fh, indent=2, default=str)
            else:
                for r in reports:
                    fh.write(render_text_report(r))
        print(f"\nReport saved to: {args.output}")

    # Excel export
    if args.xlsx and reports:
        xlsx_path = args.xlsx if args.xlsx.endswith(".xlsx") else args.xlsx + ".xlsx"
        render_xlsx_report(reports, xlsx_path)

    # Final summary
    if reports:
        print(f"\n{'=' * 72}")
        print(f"  OVERALL SUMMARY")
        print(f"{'=' * 72}")
        for r in reports:
            status_color = Colors.severity(r.overall_status)
            print(f"  {r.hostname or r.ip:30s}  {status_color}  "
                  f"({r.critical_count}C / {r.warning_count}W)")
        print(f"{'=' * 72}\n")


if __name__ == "__main__":
    main()
